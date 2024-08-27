from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment


class CXlAutofit():
    # 生成列名字典，只是为了方便修改列宽时指定列，key:数字，从1开始；value:列名，从A开始
    def get_num_colnum_dict(self):
        '''
        :return: 返回字典：{1:'A', 2:'B', ...... , 52:'AZ'}
        '''
        num_str_dict = {}
        A_Z = [chr(a) for a in range(ord('A'), ord('Z') + 1)]
        AA_AZ = ['A' + chr(a) for a in range(ord('A'), ord('Z') + 1)]
        A_AZ = A_Z + AA_AZ
        for i in A_AZ:
            num_str_dict[A_AZ.index(i) + 1] = i
        return num_str_dict

    # 自适应列宽
    def style_excel(self, excel_name: str, sheet_name: str, wb=None, bio=None):
        '''
        :param sheet_name:  excel中的sheet名
        :return:
        '''
        # 打开excel
        if wb is None:
            wb = load_workbook(excel_name)
        # 选择对应的sheet
        sheet = wb[sheet_name]
        # 获取最大行数与最大列数
        max_column = sheet.max_column
        max_row = sheet.max_row

        # 将每一列，单元格列宽最大的列宽值存到字典里，key:列的序号从1开始(与字典num_str_dic中的key对应)；value:列宽的值
        max_column_dict = {}

        # 生成列名字典，只是为了方便修改列宽时指定列，key:数字，从1开始；value:列名，从A开始
        num_str_dict = self.get_num_colnum_dict()

        # 遍历全部列
        for i in range(1, max_column + 1):
            # 遍历每一列的全部行
            for j in range(1, max_row + 1):
                column = 0
                # 获取j行i列的值
                sheet_value = sheet.cell(row=j, column=i).value
                # 通过列表生成式生成字符列表，将当前获取到的单元格的str值的每一个字符放在一个列表中（列表中一个元素是一个字符）
                sheet_value_list = [k for k in str(sheet_value)]
                # 遍历当前单元格的字符列表
                for v in sheet_value_list:
                    # 判定长度，一个数字或一个字母，单元格列宽+=1.1，其它+=2.2（长度可根据需要自行修改，经测试一个字母的列宽长度大概为1）
                    if v.isdigit() == True or v.isalpha() == True:
                        column += 1.1
                    else:
                        column += 2.2
                # 当前单元格列宽与字典中的对比，大于字典中的列宽值则将字典更新。如果字典没有这个key，抛出异常并将值添加到字典中
                try:
                    if column > max_column_dict[i]:
                        max_column_dict[i] = column
                except Exception as e:
                    max_column_dict[i] = column
        # 此时max_column_dict字典中已存有当前sheet的所有列的最大列宽值，直接遍历字典修改列宽
        for key, value in max_column_dict.items():
            sheet.column_dimensions[num_str_dict[key]].width = value
        # 保存
        if bio is None:
            wb.save(excel_name)
        else:
            wb.save(bio)
            bio.seek(0)


def gen_template(header, file_name, data=None):
    if data is None:
        data = []
    wb = Workbook()
    ws = wb.active
    ws.title = file_name
    if isinstance(header[0], str): gen_template_head_one_row(ws, header)
    if isinstance(header[0], tuple): gen_template_head_multi_row(ws, header)
    # 写入数据
    if data != []:
        start = calculate_header_rows(header)
        gen_template_with_data(ws, start, data)
    bio = BytesIO()
    # wb.save(bio)
    # bio.seek(0)

    Entity = CXlAutofit()
    Entity.style_excel('', file_name, wb=wb, bio=bio)

    return bio


def gen_template_head_one_row(ws, header):
    for i, c in enumerate(header): ws.cell(row=1, column=i + 1).value = c


def gen_template_head_multi_row(ws, header):
    x = 1
    y = 1
    for i, c in enumerate(header):
        if len(c) == 4:
            value, w, h, child = c
            merge_cells(ws, x, y, w, h, value)
            deal_child(ws, x, y + h, child)

        else:
            value, w, h = c
            merge_cells(ws, x, y, w, h, value)
        x = x + w


def deal_child(ws, x, y, child):
    for i, c in enumerate(child):
        if len(c) == 4:
            value, w, h, child = c
            merge_cells(ws, x, y, w, h, value)
            deal_child(ws, x, y + h, child)
        else:
            value, w, h = c
            merge_cells(ws, x, y, w, h, value)
        x = x + w


def merge_cells(ws, x, y, w, h, value):
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.merge_cells(start_row=y, start_column=x, end_row=y + h - 1, end_column=x + w - 1)
    ws.cell(row=y, column=x).value = value
    ws.cell(row=y, column=x).alignment = align


def calculate_header_rows(header):
    if isinstance(header[0], str):
        start = 2
    if isinstance(header[0], tuple):
        head_hight_list = []

        def dfs_head_hight(head, high=0):
            high = high + head[2]
            if len(head) == 3:
                head_hight_list.append(high)
            else:
                for h in head[3]:
                    dfs_head_hight(h, high)

        for head in header: dfs_head_hight(head)
        start = max(head_hight_list)
    return start


def gen_template_with_data(ws, start, data):
    for row_index, row in enumerate(data):
        if not type(row) == list: row = row.list()
        for col_index, col in enumerate(row):
            ws.cell(row=start + row_index, column=col_index + 1).value = col


if __name__ == '__main__':
    # header [(col_name,width,height)]
    header = [
        ("机构号", 1, 3),
        ("机构名", 1, 3),
        ("员工号", 1, 3),
        ("员工名", 1, 3),
        ("A部门", 3, 1, (
            ("综合1", 2, 1, (("薪酬1", 1, 1), ("薪酬2", 1, 1))),
            ("合计", 1, 2)
        )
         ),
    ]
